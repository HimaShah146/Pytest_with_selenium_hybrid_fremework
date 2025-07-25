import openpyxl

def get_data_by_scenario_login(file_path, sheet_name, scenario_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    data = []

    # Get column headers (first row)
    headers = [cell.value for cell in sheet[1]]

    # Get column indexes
    scenario_index = headers.index("Scenario")
    email_index = headers.index("Email")
    password_index = headers.index("Password")

    # Iterate over rows
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[scenario_index] == scenario_name:
            email = row[email_index]
            password = row[password_index]
            data.append((email, password))  # Return tuple

    return data

def get_data_by_scenario_register(file_path, sheet_name, scenario_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    data = []

    # Get column headers (first row)
    headers = [cell.value for cell in sheet[1]]

    # Get column indexes
    scenario_index = headers.index("Scenario")
    fname_index = headers.index("fname")
    lname_index = headers.index("lname")
    phone_index = headers.index("email")
    email_index = headers.index("phone")
    password_index = headers.index("passsword")
    confirm_password_index = headers.index("confirm_password")
    checkbox_index = headers.index("privacy_checkbox")

    # Iterate over rows
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[scenario_index] == scenario_name:
            fname = row[fname_index]
            lname = row[lname_index]
            phone = row[phone_index]
            email = row[email_index]
            password = row[password_index]
            confirm_password = row[confirm_password_index]
            checkbox = row[checkbox_index]

            data.append((fname, lname, phone, email, password, confirm_password, checkbox))

    return data